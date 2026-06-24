from pathlib import Path
from typing import List, Dict, Any

from openpyxl import load_workbook


class PlannerLoader:

    def __init__(self, planner_file: str):
        self.planner_file = Path(planner_file)

    def load(
        self,
        sheet_name: str = "MarketingPlan"
    ) -> List[Dict[str, Any]]:

        wb = load_workbook(
            self.planner_file,
            data_only=True
        )

        ws = wb[sheet_name]

        headers = [
            cell.value
            for cell in ws[1]
        ]

        rows = []

        for row in ws.iter_rows(
            min_row=2,
            values_only=True
        ):

            if all(v is None for v in row):
                continue

            item = dict(
                zip(headers, row)
            )

            rows.append(item)

        return rows
