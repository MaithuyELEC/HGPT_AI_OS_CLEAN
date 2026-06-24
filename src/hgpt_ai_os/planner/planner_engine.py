from openpyxl import load_workbook
from pathlib import Path

from hgpt_ai_os.planner.loader import PlannerLoader
from hgpt_ai_os.planner.validator import PlannerValidator


class PlannerEngine:

    def __init__(self, planner="planner/HGPT_MASTER_PLANNER.xlsx"):
        self.planner = Path(planner)

    def workbook(self):
        return load_workbook(self.planner)

    def marketing(self):
        wb = self.workbook()
        return wb["MarketingPlan"]

    def next_task(self):

        loader = PlannerLoader(str(self.planner))
        rows = loader.load()

        validator = PlannerValidator()
        result = validator.validate_rows(rows)

        if not result.ok:
            raise RuntimeError(
                "\n".join(result.errors)
            )

        for row_no, row in enumerate(rows, start=2):

            if str(row.get("Status", "")).upper() == "TODO":

                return {
                    "row": row_no,
                    "id": row.get("ID"),
                    "day": row.get("Day"),
                    "topic": row.get("Topic"),
                    "platform": row.get("Platform"),
                    "status": row.get("Status"),
                    "folder": row.get("Output Folder"),
                    "priority": row.get("Priority"),
                }

        return None

    def update_status(self, row, status):
        wb = self.workbook()
        ws = wb["MarketingPlan"]
        ws.cell(row=row, column=5).value = status
        wb.save(self.planner)
