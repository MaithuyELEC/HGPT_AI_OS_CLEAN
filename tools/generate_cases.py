from pathlib import Path
from openpyxl import load_workbook

TEMPLATE = Path("knowledge/CASE_TEMPLATE.md").read_text(encoding="utf-8")

wb = load_workbook("planner/STEEL_CASE_REGISTER.xlsx", data_only=True)
ws = wb["Cases"]

for row in ws.iter_rows(min_row=2, values_only=True):

    case_id, category, title = row

    if not case_id:
        continue

    out = Path("knowledge") / str(category).lower()
    out.mkdir(parents=True, exist_ok=True)

    text = TEMPLATE
    text = text.replace("Title:", f"Title: {title}")
    text = text.replace("Category:", f"Category: {category}")

    (out / f"{case_id}.md").write_text(
        text,
        encoding="utf-8"
    )

print("Knowledge Generation PASS")
